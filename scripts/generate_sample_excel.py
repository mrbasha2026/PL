#!/usr/bin/env python3
"""Generate PnL_Sample_Data.xlsx with realistic Saudi company financial data.
Includes 2025 + 2026 data for YoY comparison."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─── Color & Style Definitions ───────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
YEAR_2025_FILL = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
YEAR_2025_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
SUBHEADER_FONT = Font(name="Arial", bold=True, color="1B5E20", size=11)
SECTION_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
SECTION_FONT = Font(name="Arial", bold=True, color="1B5E20", size=11)
TOTAL_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
TOTAL_FONT = Font(name="Arial", bold=True, color="F57F17", size=11)
PROFIT_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
PROFIT_FONT = Font(name="Arial", bold=True, color="2E7D32", size=11)
NET_INCOME_FILL = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
NET_INCOME_FONT = Font(name="Arial", bold=True, color="1B5E20", size=12)
NORMAL_FONT = Font(name="Arial", size=10)
INDENT_FONT = Font(name="Arial", size=10, color="424242")
NUMBER_FMT = '#,##0'
THIN_BORDER = Border(bottom=Side(style='thin', color='E0E0E0'))
BOTTOM_BORDER = Border(bottom=Side(style='medium', color='1B5E20'))

# ─── Line Items Definition (matching pnl-types.ts) ───────────────────────────
# (nameEn, nameAr, indent, isSection, isTotal, isProfit)
LINE_ITEMS = [
    # REVENUE
    ("Revenue", "الإيرادات", 0, False, True, False),
    ("Sales Revenue", "إيرادات المبيعات", 1, False, False, False),
    ("Service Revenue", "إيرادات الخدمات", 1, False, False, False),
    ("Other Revenue", "إيرادات أخرى", 1, False, False, False),
    # COST OF SALES
    ("Cost of Goods Sold", "تكلفة البضاعة المباعة", 1, True, False, False),
    ("Raw Materials", "المواد الخام", 2, False, False, False),
    ("Direct Labor", "العمالة المباشرة", 2, False, False, False),
    ("Manufacturing Overhead", "مصروفات التصنيع غير المباشرة", 2, False, False, False),
    ("Purchases", "المشتريات", 2, False, False, False),
    # GROSS PROFIT
    ("Gross Profit", "إجمالي الربح", 0, False, False, True),
    # OPERATING EXPENSES
    ("Operating Expenses", "المصروفات التشغيلية", 0, True, True, False),
    # Selling & Marketing
    ("Selling Expenses", "مصروفات البيع والتسويق", 1, True, False, False),
    ("Sales Commissions", "عمولات المبيعات", 2, False, False, False),
    ("Advertising", "الإعلان والترويج", 2, False, False, False),
    ("Marketing Expenses", "مصروفات التسويق", 2, False, False, False),
    ("Delivery & Shipping", "التوصيل والشحن", 2, False, False, False),
    ("Customer Service", "خدمة العملاء", 2, False, False, False),
    # G&A
    ("General & Administrative", "المصروفات الإدارية والعمومية", 1, True, False, False),
    ("Salaries & Wages", "الرواتب والأجور", 2, False, False, False),
    ("Employee Benefits", "بدلات ومزايا الموظفين", 2, False, False, False),
    ("GOSI Contributions", "اشتراكات التأمينات الاجتماعية (GOSI)", 2, False, False, False),
    ("Rent Expense", "الإيجارات", 2, False, False, False),
    ("Utilities", "المرافق (كهرباء وماء وغاز)", 2, False, False, False),
    ("Telecommunications", "الاتصالات والإنترنت", 2, False, False, False),
    ("Office Supplies", "القرطاسية واللوازم المكتبية", 2, False, False, False),
    ("Professional Fees", "الأتعاب المهنية", 2, False, False, False),
    ("Travel & Entertainment", "السفر والضيافة", 2, False, False, False),
    ("Insurance Expense", "التأمين", 2, False, False, False),
    ("Maintenance & Repairs", "الصيانة والإصلاحات", 2, False, False, False),
    ("Licenses & Permits", "التراخيص والرسوم الحكومية", 2, False, False, False),
    ("Subscriptions & Software", "الاشتراكات والبرمجيات", 2, False, False, False),
    ("Bad Debts", "الديون المعدومة", 2, False, False, False),
    ("Miscellaneous Expenses", "مصروفات متنوعة", 2, False, False, False),
    # Depreciation
    ("Depreciation & Amortization", "الإهلاك والاستنفاد", 1, True, False, False),
    ("Depreciation of Buildings", "إهلاك المباني", 2, False, False, False),
    ("Depreciation of Equipment", "إهلاك المعدات والأجهزة", 2, False, False, False),
    ("Depreciation of Vehicles", "إهلاك السيارات", 2, False, False, False),
    ("Amortization of Intangibles", "استنفاد الأصول غير الملموسة", 2, False, False, False),
    # OPERATING INCOME
    ("Operating Income (EBIT)", "الدخل التشغيلي", 0, False, False, True),
    # NON-OPERATING
    ("Interest Income", "إيرادات الاستثمارات", 1, False, False, False),
    ("Finance Cost", "تكلفة التمويل", 1, False, False, False),
    ("Other Income", "إيرادات أخرى", 1, False, False, False),
    ("Other Expenses", "مصروفات أخرى", 1, False, False, False),
    # INCOME BEFORE ZAKAT
    ("Income Before Zakat", "الدخل قبل الزكاة", 0, False, False, True),
    # ZAKAT
    ("Zakat Expense", "مصروف الزكاة", 1, False, False, False),
    # NET INCOME
    ("Net Income", "صافي الدخل", 0, False, True, True),
]

# 12 months: Jan-Jun 2025 + Jan-Jun 2026
PERIODS = [
    "Jan 2025", "Feb 2025", "Mar 2025", "Apr 2025", "May 2025", "Jun 2025",
    "Jan 2026", "Feb 2026", "Mar 2026", "Apr 2026", "May 2026", "Jun 2026",
]

# ─── Helper: scale a list by a factor ────────────────────────────────────────
def scale(vals, factor):
    return [round(v * factor) for v in vals]

# ─── 2026 data (same as before) ──────────────────────────────────────────────
# Company 1: شركة النخبة التجارية — Trading, ~8M/month in 2026
nukhba_2026 = {
    "Revenue":                          [7200000, 7500000, 7800000, 8100000, 8500000, 8900000],
    "Sales Revenue":                    [6500000, 6800000, 7000000, 7300000, 7700000, 8000000],
    "Service Revenue":                  [500000, 520000, 550000, 570000, 550000, 600000],
    "Other Revenue":                    [200000, 180000, 250000, 230000, 250000, 300000],
    "Cost of Goods Sold":               [3960000, 4125000, 4290000, 4455000, 4675000, 4895000],
    "Raw Materials":                    [2160000, 2250000, 2340000, 2430000, 2550000, 2670000],
    "Direct Labor":                     [900000, 937500, 975000, 1012500, 1062500, 1113750],
    "Manufacturing Overhead":           [540000, 562500, 585000, 607500, 637500, 668250],
    "Purchases":                        [360000, 375000, 390000, 405000, 425000, 443000],
    "Gross Profit":                     [3240000, 3375000, 3510000, 3645000, 3825000, 4005000],
    "Operating Expenses":               [2520000, 2625000, 2730000, 2835000, 2975000, 3115000],
    "Selling Expenses":                 [720000, 750000, 780000, 810000, 850000, 890000],
    "Sales Commissions":                [195000, 204000, 210000, 219000, 231000, 240000],
    "Advertising":                      [144000, 150000, 156000, 162000, 170000, 178000],
    "Marketing Expenses":               [216000, 225000, 234000, 243000, 255000, 267000],
    "Delivery & Shipping":              [108000, 112500, 117000, 121500, 127500, 133500],
    "Customer Service":                 [57000, 58500, 63000, 64500, 66500, 71500],
    "General & Administrative":         [1296000, 1350000, 1404000, 1458000, 1530000, 1602000],
    "Salaries & Wages":                 [540000, 540000, 540000, 567000, 567000, 567000],
    "Employee Benefits":                [162000, 162000, 162000, 170100, 170100, 170100],
    "GOSI Contributions":               [75600, 75600, 75600, 79380, 79380, 79380],
    "Rent Expense":                     [120000, 120000, 120000, 120000, 120000, 120000],
    "Utilities":                        [45000, 42000, 40000, 48000, 54000, 58000],
    "Telecommunications":               [18000, 18000, 18000, 18000, 18000, 18000],
    "Office Supplies":                  [12000, 11500, 10000, 13000, 12500, 11000],
    "Professional Fees":                [60000, 55000, 75000, 50000, 65000, 80000],
    "Travel & Entertainment":           [35000, 30000, 28000, 45000, 50000, 65000],
    "Insurance Expense":                [24000, 24000, 24000, 24000, 24000, 24000],
    "Maintenance & Repairs":            [30000, 35000, 25000, 40000, 32000, 38000],
    "Licenses & Permits":               [15000, 10000, 8000, 12000, 9000, 8000],
    "Subscriptions & Software":         [36000, 36000, 36000, 36000, 36000, 36000],
    "Bad Debts":                        [45000, 48000, 52000, 50000, 55000, 60000],
    "Miscellaneous Expenses":           [77400, 80900, 84400, 83520, 88020, 93520],
    "Depreciation & Amortization":      [360000, 360000, 360000, 360000, 375000, 375000],
    "Depreciation of Buildings":        [80000, 80000, 80000, 80000, 80000, 80000],
    "Depreciation of Equipment":        [180000, 180000, 180000, 180000, 180000, 180000],
    "Depreciation of Vehicles":         [60000, 60000, 60000, 60000, 75000, 75000],
    "Amortization of Intangibles":      [40000, 40000, 40000, 40000, 40000, 40000],
    "Operating Income (EBIT)":          [720000, 750000, 780000, 810000, 850000, 890000],
    "Interest Income":                  [45000, 42000, 48000, 50000, 55000, 52000],
    "Finance Cost":                     [120000, 115000, 110000, 105000, 100000, 95000],
    "Other Income":                     [30000, 25000, 35000, 40000, 28000, 45000],
    "Other Expenses":                   [15000, 18000, 12000, 20000, 15000, 10000],
    "Income Before Zakat":              [660000, 684000, 741000, 775000, 818000, 882000],
    "Zakat Expense":                    [16500, 17100, 18525, 19375, 20450, 22050],
    "Net Income":                       [643500, 666900, 722475, 755625, 797550, 859950],
}

# Company 2: شركة الأفق للخدمات — Services, ~4.5M/month in 2026
ufuq_2026 = {
    "Revenue":                          [4200000, 4350000, 4500000, 4650000, 4800000, 4950000],
    "Sales Revenue":                    [1200000, 1250000, 1300000, 1350000, 1400000, 1450000],
    "Service Revenue":                  [2800000, 2900000, 3000000, 3100000, 3200000, 3300000],
    "Other Revenue":                    [200000, 200000, 200000, 200000, 200000, 200000],
    "Cost of Goods Sold":               [1680000, 1740000, 1800000, 1860000, 1920000, 1980000],
    "Raw Materials":                    [420000, 435000, 450000, 465000, 480000, 495000],
    "Direct Labor":                     [630000, 652500, 675000, 697500, 720000, 742500],
    "Manufacturing Overhead":           [336000, 348000, 360000, 372000, 384000, 396000],
    "Purchases":                        [294000, 304500, 315000, 325500, 336000, 346500],
    "Gross Profit":                     [2520000, 2610000, 2700000, 2790000, 2880000, 2970000],
    "Operating Expenses":               [1890000, 1957500, 2025000, 2092500, 2160000, 2227500],
    "Selling Expenses":                 [378000, 391500, 405000, 418500, 432000, 445500],
    "Sales Commissions":                [126000, 130500, 135000, 139500, 144000, 148500],
    "Advertising":                      [75600, 78300, 81000, 83700, 86400, 89100],
    "Marketing Expenses":               [88200, 91350, 94500, 97650, 100800, 103950],
    "Delivery & Shipping":              [50400, 52200, 54000, 55800, 57600, 59400],
    "Customer Service":                 [37800, 39150, 40500, 41850, 43200, 44550],
    "General & Administrative":         [1134000, 1174500, 1215000, 1255500, 1296000, 1336500],
    "Salaries & Wages":                 [540000, 540000, 540000, 567000, 567000, 567000],
    "Employee Benefits":                [135000, 135000, 135000, 141750, 141750, 141750],
    "GOSI Contributions":               [67500, 67500, 67500, 70875, 70875, 70875],
    "Rent Expense":                     [90000, 90000, 90000, 90000, 90000, 90000],
    "Utilities":                        [32000, 30000, 28000, 35000, 38000, 42000],
    "Telecommunications":               [15000, 15000, 15000, 15000, 15000, 15000],
    "Office Supplies":                  [8000, 7500, 7000, 8500, 8000, 7500],
    "Professional Fees":                [45000, 50000, 40000, 55000, 60000, 45000],
    "Travel & Entertainment":           [28000, 25000, 22000, 35000, 40000, 50000],
    "Insurance Expense":                [18000, 18000, 18000, 18000, 18000, 18000],
    "Maintenance & Repairs":            [20000, 25000, 18000, 30000, 22000, 28000],
    "Licenses & Permits":               [10000, 8000, 6000, 10000, 7000, 5000],
    "Subscriptions & Software":         [60000, 60000, 60000, 60000, 60000, 60000],
    "Bad Debts":                        [28000, 30000, 35000, 32000, 38000, 40000],
    "Miscellaneous Expenses":           [51500, 52500, 54000, 53875, 55375, 58375],
    "Depreciation & Amortization":      [252000, 252000, 252000, 252000, 252000, 252000],
    "Depreciation of Buildings":        [50000, 50000, 50000, 50000, 50000, 50000],
    "Depreciation of Equipment":        [120000, 120000, 120000, 120000, 120000, 120000],
    "Depreciation of Vehicles":         [45000, 45000, 45000, 45000, 45000, 45000],
    "Amortization of Intangibles":      [37000, 37000, 37000, 37000, 37000, 37000],
    "Operating Income (EBIT)":          [630000, 652500, 675000, 697500, 720000, 742500],
    "Interest Income":                  [25000, 23000, 27000, 28000, 30000, 29000],
    "Finance Cost":                     [80000, 78000, 75000, 72000, 70000, 68000],
    "Other Income":                     [15000, 18000, 12000, 20000, 16000, 22000],
    "Other Expenses":                   [10000, 12000, 8000, 14000, 10000, 8000],
    "Income Before Zakat":              [580000, 603500, 631000, 659500, 686000, 717500],
    "Zakat Expense":                    [14500, 15088, 15775, 16488, 17150, 17938],
    "Net Income":                       [565500, 588413, 615225, 643013, 668850, 699563],
}

# Company 3: شركة البناء الحديث — Construction, ~12M/month in 2026
bina_2026 = {
    "Revenue":                          [11500000, 12000000, 11800000, 12500000, 13000000, 13500000],
    "Sales Revenue":                    [10500000, 11000000, 10800000, 11500000, 12000000, 12500000],
    "Service Revenue":                  [800000, 800000, 800000, 800000, 800000, 800000],
    "Other Revenue":                    [200000, 200000, 200000, 200000, 200000, 200000],
    "Cost of Goods Sold":               [6900000, 7200000, 7080000, 7500000, 7800000, 8100000],
    "Raw Materials":                    [4600000, 4800000, 4720000, 5000000, 5200000, 5400000],
    "Direct Labor":                     [1380000, 1440000, 1416000, 1500000, 1560000, 1620000],
    "Manufacturing Overhead":           [575000, 600000, 590000, 625000, 650000, 675000],
    "Purchases":                        [345000, 360000, 354000, 375000, 390000, 405000],
    "Gross Profit":                     [4600000, 4800000, 4720000, 5000000, 5200000, 5400000],
    "Operating Expenses":               [3220000, 3360000, 3304000, 3500000, 3640000, 3780000],
    "Selling Expenses":                 [805000, 840000, 826000, 875000, 910000, 945000],
    "Sales Commissions":                [241500, 252000, 247800, 262500, 273000, 283500],
    "Advertising":                      [161000, 168000, 165200, 175000, 182000, 189000],
    "Marketing Expenses":               [241500, 252000, 247800, 262500, 273000, 283500],
    "Delivery & Shipping":              [120750, 126000, 123900, 131250, 136500, 141750],
    "Customer Service":                 [40250, 42000, 41300, 43750, 45500, 47250],
    "General & Administrative":         [1840000, 1920000, 1888000, 2000000, 2080000, 2160000],
    "Salaries & Wages":                 [750000, 750000, 750000, 787500, 787500, 787500],
    "Employee Benefits":                [225000, 225000, 225000, 236250, 236250, 236250],
    "GOSI Contributions":               [105000, 105000, 105000, 110250, 110250, 110250],
    "Rent Expense":                     [200000, 200000, 200000, 200000, 200000, 200000],
    "Utilities":                        [85000, 80000, 75000, 95000, 110000, 120000],
    "Telecommunications":               [25000, 25000, 25000, 25000, 25000, 25000],
    "Office Supplies":                  [18000, 17000, 16000, 20000, 19000, 18000],
    "Professional Fees":                [90000, 85000, 95000, 80000, 100000, 110000],
    "Travel & Entertainment":           [50000, 45000, 42000, 60000, 65000, 75000],
    "Insurance Expense":                [35000, 35000, 35000, 35000, 35000, 35000],
    "Maintenance & Repairs":            [55000, 60000, 48000, 70000, 55000, 65000],
    "Licenses & Permits":               [20000, 15000, 12000, 18000, 14000, 10000],
    "Subscriptions & Software":         [40000, 40000, 40000, 40000, 40000, 40000],
    "Bad Debts":                        [70000, 75000, 80000, 72000, 85000, 90000],
    "Miscellaneous Expenses":           [82000, 85000, 81000, 88000, 91500, 96000],
    "Depreciation & Amortization":      [575000, 600000, 590000, 625000, 650000, 675000],
    "Depreciation of Buildings":        [150000, 150000, 150000, 150000, 150000, 150000],
    "Depreciation of Equipment":        [300000, 300000, 300000, 300000, 300000, 300000],
    "Depreciation of Vehicles":         [75000, 75000, 75000, 75000, 75000, 75000],
    "Amortization of Intangibles":      [50000, 75000, 65000, 100000, 125000, 150000],
    "Operating Income (EBIT)":          [1380000, 1440000, 1416000, 1500000, 1560000, 1620000],
    "Interest Income":                  [60000, 55000, 65000, 70000, 75000, 70000],
    "Finance Cost":                     [180000, 175000, 170000, 165000, 160000, 155000],
    "Other Income":                     [40000, 35000, 45000, 50000, 30000, 55000],
    "Other Expenses":                   [25000, 30000, 20000, 35000, 25000, 15000],
    "Income Before Zakat":              [1275000, 1325000, 1336000, 1420000, 1480000, 1575000],
    "Zakat Expense":                    [31875, 33125, 33400, 35500, 37000, 39375],
    "Net Income":                       [1243125, 1291875, 1302600, 1384500, 1443000, 1535625],
}

# ─── 2025 data: ~12-18% lower than 2026 (YoY growth story) ──────────────────
# Fixed costs (rent, salaries) stay similar; variable costs scale with revenue
# Finance costs are higher in 2025 (paying down debt)

def make_2025(data_2026, revenue_factor=0.85, variable_factor=0.85, fixed_factor=0.95):
    """Generate 2025 data from 2026 with realistic YoY differences.
    - Revenue & variable costs: scaled by revenue_factor
    - Fixed costs (salaries, rent, insurance, depreciation): scaled by fixed_factor  
    - Finance cost: higher in 2025 (more debt)
    - Bad debts: higher in 2025 (weaker collection)
    """
    # Items that are mostly fixed (don't scale much with revenue)
    FIXED_ITEMS = {
        "Salaries & Wages", "Employee Benefits", "GOSI Contributions",
        "Rent Expense", "Insurance Expense", "Telecommunications",
        "Subscriptions & Software", "Depreciation of Buildings",
        "Depreciation of Equipment", "Depreciation of Vehicles",
        "Amortization of Intangibles",
    }
    # Items that should be HIGHER in 2025 (worse performance)
    HIGHER_ITEMS = {
        "Finance Cost", "Bad Debts",
    }
    # Computed/subtotal items that need special handling
    SUBTOTAL_ITEMS = {
        "Revenue", "Cost of Goods Sold", "Gross Profit", "Operating Expenses",
        "Selling Expenses", "General & Administrative",
        "Depreciation & Amortization",
        "Operating Income (EBIT)", "Income Before Zakat", "Net Income",
    }

    result = {}
    for key, vals in data_2026.items():
        if key in FIXED_ITEMS:
            result[key] = scale(vals, fixed_factor)
        elif key in HIGHER_ITEMS:
            # ~15-25% higher in 2025
            result[key] = scale(vals, 1.20)
        elif key in SUBTOTAL_ITEMS:
            # These will be recalculated, but for simplicity use revenue_factor
            # The app recalculates them anyway
            if key in ("Operating Income (EBIT)", "Income Before Zakat", "Net Income", "Gross Profit"):
                result[key] = scale(vals, revenue_factor * 0.9)  # Profits grow faster
            else:
                result[key] = scale(vals, revenue_factor)
        else:
            result[key] = scale(vals, variable_factor)
    return result

nukhba_2025 = make_2025(nukhba_2026, revenue_factor=0.86, variable_factor=0.86, fixed_factor=0.95)
ufuq_2025   = make_2025(ufuq_2026,   revenue_factor=0.88, variable_factor=0.88, fixed_factor=0.96)
bina_2025   = make_2025(bina_2026,   revenue_factor=0.84, variable_factor=0.84, fixed_factor=0.94)

# Merge 2025 + 2026 into single 12-month arrays
def merge_years(d2025, d2026):
    return {k: d2025[k] + d2026[k] for k in d2026}

al_nukhba = merge_years(nukhba_2025, nukhba_2026)
al_ufuq   = merge_years(ufuq_2025, ufuq_2026)
al_bina   = merge_years(bina_2025, bina_2026)


def write_company_sheet(wb, sheet_name, company_data, currency="SAR"):
    """Write a company sheet with formatted P&L data for 12 months."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = "1B5E20"

    num_cols = len(PERIODS) + 1  # label + 12 periods

    # Column widths
    ws.column_dimensions['A'].width = 50
    for col_idx in range(2, num_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # Row 1: Header (Periods) — 2025 in blue, 2026 in green
    cell = ws.cell(row=1, column=1, value="البند Line Item")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for col_idx, period in enumerate(PERIODS, 2):
        cell = ws.cell(row=1, column=col_idx, value=period)
        if "2025" in period:
            cell.font = YEAR_2025_FONT
            cell.fill = YEAR_2025_FILL
        else:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Row 2: Currency
    cell = ws.cell(row=2, column=1, value="العملة Currency")
    cell.font = SUBHEADER_FONT
    cell.fill = SUBHEADER_FILL
    for col_idx in range(2, num_cols + 1):
        cell = ws.cell(row=2, column=col_idx, value=currency)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    # Freeze panes
    ws.freeze_panes = 'B3'

    # Data rows
    row = 3
    for (nameEn, nameAr, indent, isSection, isTotal, isProfit) in LINE_ITEMS:
        label = f"{'  ' * indent}{nameAr} - {nameEn}"
        cell_label = ws.cell(row=row, column=1, value=label)

        if isProfit and isTotal:
            label_font = NET_INCOME_FONT
            label_fill = NET_INCOME_FILL
        elif isProfit:
            label_font = PROFIT_FONT
            label_fill = PROFIT_FILL
        elif isTotal:
            label_font = TOTAL_FONT
            label_fill = TOTAL_FILL
        elif isSection:
            label_font = SECTION_FONT
            label_fill = SECTION_FILL
        elif indent > 0:
            label_font = INDENT_FONT
            label_fill = None
        else:
            label_font = NORMAL_FONT
            label_fill = None

        cell_label.font = label_font
        if label_fill:
            cell_label.fill = label_fill

        values = company_data.get(nameEn, [0] * len(PERIODS))
        for col_idx, val in enumerate(values, 2):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.number_format = NUMBER_FMT
            cell.alignment = Alignment(horizontal='center')
            cell.font = label_font
            if label_fill:
                cell.fill = label_fill
            if isProfit or isTotal:
                cell.border = BOTTOM_BORDER
            else:
                cell.border = THIN_BORDER

        if isProfit or isTotal:
            cell_label.border = BOTTOM_BORDER
        else:
            cell_label.border = THIN_BORDER

        row += 1


# ─── Instructions Sheet ──────────────────────────────────────────────────────
ws_instr = wb.active
ws_instr.title = "تعليمات Instructions"
ws_instr.column_dimensions['A'].width = 90
ws_instr.sheet_properties.tabColor = "1B5E20"

instructions = [
    ("تعليمات - Instructions", Font(name="Arial", bold=True, size=14, color="1B5E20")),
    ("", None),
    ("العربية:", Font(name="Arial", bold=True, size=11, color="1B5E20")),
    ("  هذا الملف يحتوي على بيانات نموذجية لثلاث شركات سعودية لسنتين (2025 + 2026)", None),
    ("  كل ورقة تمثل شركة مستقلة ببيانات 12 شهر (يناير - يونيو 2025 + يناير - يونيو 2026)", None),
    ("  الأعمدة الزرقاء = 2025 | الأعمدة الخضراء = 2026 — لدعم المقارنة السنوية", None),
    ("  يمكنك تعديل الأرقام أو إضافة شركات أخرى بأوراق جديدة", None),
    ("  اسم الورقة = اسم الشركة", None),
    ("  الصف 1: الفترات المالية  |  الصف 2: العملة  |  عمود A = اسم البند", None),
    ("", None),
    ("  ملاحظة: لا توجد ضريبة دخل — نظام الزكاة الشرعية ينطبق", Font(name="Arial", bold=True, size=10, color="D32F2F")),
    ("  البنود باللون الأخضر = أرباح | باللون الأصفر = إجماليات | بالأخضر الفاتح = أقسام", None),
    ("", None),
    ("English:", Font(name="Arial", bold=True, size=11, color="1B5E20")),
    ("  This file contains sample data for three Saudi companies for 2 years (2025 + 2026)", None),
    ("  Each sheet = one company with 12 months (Jan-Jun 2025 + Jan-Jun 2026)", None),
    ("  Blue columns = 2025 | Green columns = 2026 — enables year-over-year comparison", None),
    ("  You can edit numbers or add more companies in new sheets", None),
    ("  Sheet name = Company name", None),
    ("  Row 1: Periods  |  Row 2: Currency  |  Column A = Item name", None),
    ("", None),
    ("  Note: No income tax — Islamic Zakat system applies", Font(name="Arial", bold=True, size=10, color="D32F2F")),
    ("  Green cells = Profits | Yellow cells = Totals | Light green = Sections", None),
    ("", None),
    ("أوراق الحركات / Movement Sheets:", Font(name="Arial", bold=True, size=11, color="0D47A1")),
    ("  أوراق 'حركات X' تحتوي على القيم + نسبة التغير الدوري بين كل فترة والتي قبلها", None),
    ("  اللون الأخضر = زيادة > 2% | اللون الأحمر = انخفاض > 2% | الرمادي = تغير طفيف", None),
    ("  المتوسط = متوسط نسبة التغير خلال كل الفترات", None),
    ("  هذه البيانات تحسب تلقائياً في الموقع عند اختيار تبويب 'حركات البنود'", None),
    ("", None),
    ("  Movement sheets ('حركات X') contain values + period-over-period % change", None),
    ("  Green = increase > 2% | Red = decrease > 2% | Grey = minor change", None),
    ("  Average = mean % change across all periods", None),
    ("  This data is auto-calculated on the website in the 'Line Item Movements' tab", None),
    ("", None),
    ("الشركات النموذجية / Sample Companies:", Font(name="Arial", bold=True, size=11, color="1B5E20")),
    ("  1. شركة النخبة التجارية — تجارية (~7M→8M ريال/شهر) — Trading (~14% YoY growth)", None),
    ("  2. شركة الأفق للخدمات — خدماتية (~4M→4.5M ريال/شهر) — Services (~12% YoY growth)", None),
    ("  3. شركة البناء الحديث — صناعية/إنشائية (~10M→12M ريال/شهر) — Construction (~19% YoY growth)", None),
]

for idx, (text, font) in enumerate(instructions, 1):
    cell = ws_instr.cell(row=idx, column=1, value=text)
    if font:
        cell.font = font
    else:
        cell.font = Font(name="Arial", size=10)


# ─── Movement Analysis Sheet ────────────────────────────────────────────────
# Shows period-over-period % change for all line items per company

MOVEMENT_FILL_POS = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # green for positive
MOVEMENT_FILL_NEG = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")  # red for negative
MOVEMENT_FILL_ZERO = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # grey for zero
MOVEMENT_FONT_POS = Font(name="Arial", size=10, color="2E7D32", bold=True)
MOVEMENT_FONT_NEG = Font(name="Arial", size=10, color="C62828", bold=True)
MOVEMENT_FONT_ZERO = Font(name="Arial", size=10, color="757575")
PCT_FMT = '0.0%'

def write_movement_sheet(wb, sheet_name, company_data, currency="SAR"):
    """Write a movement analysis sheet showing period-over-period % changes."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = "0D47A1"  # Blue tab for movement sheets

    num_value_cols = len(PERIODS)
    num_change_cols = num_value_cols - 1  # one fewer change column
    total_cols = 1 + num_value_cols + 2 + num_change_cols + 2  # label + values + (total, avg) + changes + (avg change, total change)

    # Column widths
    ws.column_dimensions['A'].width = 50
    for col_idx in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # ─── Section 1: Values ────────────────────────────────
    # Header row
    col = 1
    cell = ws.cell(row=1, column=col, value="البند Line Item")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')

    col = 2
    for period in PERIODS:
        cell = ws.cell(row=1, column=col, value=period)
        if "2025" in period:
            cell.font = YEAR_2025_FONT
            cell.fill = YEAR_2025_FILL
        else:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        col += 1

    # Total + Average
    cell = ws.cell(row=1, column=col, value="الإجمالي")
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color="F57F17", end_color="F57F17", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    col += 1
    cell = ws.cell(row=1, column=col, value="المتوسط")
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color="F57F17", end_color="F57F17", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    col += 1

    # ─── Section 2: % Changes ────────────────────────────────
    # Separator column
    cell = ws.cell(row=1, column=col, value="")
    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    col += 1

    # Change headers
    cell = ws.cell(row=1, column=col, value="نسبة التغير الدوري %")
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    # Merge across change columns
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + num_change_cols + 1)

    # Sub-header for change columns
    col = col  # stay at same col for row 2
    change_start_col = col
    for i in range(1, len(PERIODS)):
        cell = ws.cell(row=2, column=col, value=f"{PERIODS[i]}\nvs\n{PERIODS[i-1]}")
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        col += 1

    # Average change + Total change
    cell = ws.cell(row=2, column=col, value="متوسط\nالتغير %")
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    cell.fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    col += 1

    # Row 2: sub-headers for value section
    for c in range(1, change_start_col):
        if ws.cell(row=2, column=c).value is None:
            cell = ws.cell(row=2, column=c, value="")
            cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    # Freeze panes
    ws.freeze_panes = 'B3'
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 50

    # Data rows
    row = 3
    for (nameEn, nameAr, indent, isSection, isTotal, isProfit) in LINE_ITEMS:
        label = f"{'  ' * indent}{nameAr} - {nameEn}"
        cell_label = ws.cell(row=row, column=1, value=label)

        if isProfit and isTotal:
            label_font = NET_INCOME_FONT
            label_fill = NET_INCOME_FILL
        elif isProfit:
            label_font = PROFIT_FONT
            label_fill = PROFIT_FILL
        elif isTotal:
            label_font = TOTAL_FONT
            label_fill = TOTAL_FILL
        elif isSection:
            label_font = SECTION_FONT
            label_fill = SECTION_FILL
        elif indent > 0:
            label_font = INDENT_FONT
            label_fill = None
        else:
            label_font = NORMAL_FONT
            label_fill = None

        cell_label.font = label_font
        if label_fill:
            cell_label.fill = label_fill

        values = company_data.get(nameEn, [0] * len(PERIODS))

        # Write values
        col = 2
        total_val = 0
        for val in values:
            cell = ws.cell(row=row, column=col, value=val)
            cell.number_format = NUMBER_FMT
            cell.alignment = Alignment(horizontal='center')
            cell.font = label_font
            if label_fill:
                cell.fill = label_fill
            total_val += val
            col += 1

        # Total + Average
        cell = ws.cell(row=row, column=col, value=total_val)
        cell.number_format = NUMBER_FMT
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        col += 1

        avg_val = total_val / len(values) if values else 0
        cell = ws.cell(row=row, column=col, value=round(avg_val))
        cell.number_format = NUMBER_FMT
        cell.font = Font(name="Arial", size=10)
        cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        col += 1

        # Separator
        cell = ws.cell(row=row, column=col, value="")
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        col += 1

        # % Changes
        pct_changes = []
        for i in range(1, len(values)):
            prev = values[i - 1]
            curr = values[i]
            if prev != 0:
                pct = (curr - prev) / abs(prev)
            else:
                pct = 0
            pct_changes.append(pct)

            cell = ws.cell(row=row, column=col, value=pct)
            cell.number_format = PCT_FMT
            cell.alignment = Alignment(horizontal='center')

            # Conditional formatting by color
            if pct > 0.02:
                cell.fill = MOVEMENT_FILL_POS
                cell.font = MOVEMENT_FONT_POS
            elif pct < -0.02:
                cell.fill = MOVEMENT_FILL_NEG
                cell.font = MOVEMENT_FONT_NEG
            else:
                cell.fill = MOVEMENT_FILL_ZERO
                cell.font = MOVEMENT_FONT_ZERO
            col += 1

        # Average change
        if pct_changes:
            avg_pct = sum(pct_changes) / len(pct_changes)
            cell = ws.cell(row=row, column=col, value=avg_pct)
            cell.number_format = PCT_FMT
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(name="Arial", bold=True, size=10,
                           color="2E7D32" if avg_pct > 0 else "C62828" if avg_pct < 0 else "757575")
        col += 1

        row += 1

    return ws


# ─── Write Company Sheets ────────────────────────────────────────────────────
write_company_sheet(wb, "النخبة التجارية", al_nukhba)
write_company_sheet(wb, "الأفق للخدمات", al_ufuq)
write_company_sheet(wb, "البناء الحديث", al_bina)

# ─── Write Movement Sheets ───────────────────────────────────────────────────
write_movement_sheet(wb, "حركات النخبة", al_nukhba)
write_movement_sheet(wb, "حركات الأفق", al_ufuq)
write_movement_sheet(wb, "حركات البناء", al_bina)

# ─── Save ─────────────────────────────────────────────────────────────────────
output_path = "/home/z/my-project/download/PnL_Sample_Data.xlsx"
wb.save(output_path)
print(f"✅ File saved to: {output_path}")
print(f"   Sheets: {wb.sheetnames}")
print(f"   Companies: 3 (+ 3 movement sheets)")
print(f"   Periods per company: 12 (Jan-Jun 2025 + Jan-Jun 2026)")
print(f"   YoY growth: النخبة ~14% | الأفق ~12% | البناء ~19%")
